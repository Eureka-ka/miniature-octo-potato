# -*- coding: utf-8 -*-
"""问题1：数学建模论文质量综合评价（优化版）
沿用 AHP + 模糊综合评价骨架，改进点：
1) AHP权重由 AHP_A.m 的判断矩阵复算并给出 CR；
2) 二级指标采用“参照锚定”的绝对化得分（覆盖率类以0.8为达标线，
   密度/数量类以样本80分位为参照，反向指标取补），映射为连续隶属度；
3) 分级采用综合得分法（B·[90,75,60,45,25]，阈值80/65/50/35）；
4) 输出每个二级指标得分（供问题3定位薄弱项）；
5) 图像型论文（如25.pdf）标记并排除。
"""
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import common as C
import quantify

IND_ORDER = [f"u{i}{j}" for i in range(1, 5) for j in range(1, 5)]
KIND = quantify.KIND

def dim_inds(dim):
    return [ind for ind in IND_ORDER if C.INDICATOR_META[ind][1] == dim]

def abs_score(raw_arr, kind):
    """把某指标的原始值序列映射为绝对化得分 s∈[0,1]（可复用于问题3）。"""
    arr = np.asarray(raw_arr, dtype=float)
    if kind == "ratio":
        return np.minimum(1.0, arr / 0.6)          # 覆盖率达标线 60%（描述性词典覆盖率）
    if kind == "density":
        ref = float(np.percentile(arr, 80))
        if ref <= 0:
            return np.full_like(arr, 0.5)
        return np.minimum(1.0, arr / ref)          # 密度以样本80分位为参照
    if kind == "low":
        ref = float(np.percentile(arr, 80))
        if ref <= 0:
            return np.ones_like(arr)               # 未检出缺陷视为满分
        return 1.0 - np.minimum(1.0, arr / ref)    # 越低越好
    if kind == "count":
        ref = max(8.0, float(np.percentile(arr, 80)))
        return np.minimum(1.0, arr / ref)          # 引用条数达标线≥8
    raise ValueError(kind)

def run():
    files = sorted(f for f in os.listdir(C.ATT1) if f.lower().endswith(".pdf"))
    rows = []
    for f in files:
        path = os.path.join(C.ATT1, f)
        pages, full = C.extract_pdf(path)
        base = {"name": f, "image_only": C.is_image_only(full)}
        if base["image_only"]:
            base["topic"] = "图像型论文"
            rows.append(base)
            continue
        t = C.analyze_text(pages, full)
        topic = C.classify_topic(full)
        raw = quantify.quantify_all(t, topic)
        rows.append(dict(base, t=t, topic=topic, raw=raw))

    usable = [r for r in rows if not r.get("image_only") and r.get("topic") != "人工审核"]

    # ---- 参照锚定绝对化得分 ----
    refs = {}
    raw_by_ind = {ind: np.array([r["raw"][ind] for r in usable]) for ind in IND_ORDER}
    ref_vals = C.compute_refs(raw_by_ind, KIND)
    for ind in IND_ORDER:
        s = C.apply_refs(raw_by_ind[ind], KIND[ind], ref_vals[ind])
        for r, sv in zip(usable, s):
            r.setdefault("s", {})[ind] = float(sv)
        name, dim, direc = C.INDICATOR_META[ind]
        refs[ind] = {"name": name, "dim": dim, "kind": KIND[ind],
                     "ref": ref_vals[ind],
                     "min": float(raw_by_ind[ind].min()), "max": float(raw_by_ind[ind].max())}

    # ---- 等级阈值校准（参照样本分位 p10..p90）----
    s_by_ind = {ind: np.array([r["s"][ind] for r in usable]) for ind in IND_ORDER}
    centers, widths = C.calibrate_centers(s_by_ind)

    # ---- AHP权重 ----
    weights = C.compute_all_weights()
    A = {"A": weights["M(一级)"]["w"]}
    for i in range(1, 5):
        key = [k for k in weights if k.startswith(f"M{i}(")][0]
        A[f"A{i}"] = weights[key]["w"]

    # ---- 模糊综合评价 ----
    scorer = C.FCEScorer(weights, centers, widths)
    ind_order = {dim: dim_inds(dim) for dim in C.U_ORDER}
    for r in usable:
        norm_scores = {dim: [r["s"][ind] for ind in dim_inds(dim)] for dim in C.U_ORDER}
        B, score = scorer.score(norm_scores, ind_order)
        dim_scores = {}
        for j, dim in enumerate(C.U_ORDER):
            Rj = np.array([scorer.membership(ind, r["s"][ind]) for ind in dim_inds(dim)])
            dim_scores[dim] = float((A[f"A{j+1}"] @ Rj) @ C.GRADE_SCORES)
        r["B"] = B
        r["score"] = float(score)
        r["dim_scores"] = dim_scores
    # 相对分位定级：按批次得分分布拉开五级（增强区分度）
    bands = C.percentile_bands([r["score"] for r in usable])
    for r in usable:
        r["level"] = C.level_by_band(r["score"], bands)

    # ---- 导出 ----
    export_excel(rows, usable, weights, os.path.join(C.OUT, "problem1_results.xlsx"))
    export_csv(usable, os.path.join(C.OUT, "problem1_indicators.csv"))
    C.save_json({"refs": refs,
                 "bands": bands,
                 "centers": {k: v.tolist() for k, v in centers.items()},
                 "widths": {k: v for k, v in widths.items()},
                 "weights": {k: {"w": v["w"].tolist(), "CR": v["CR"], "CI": v["CI"], "lam": v["lam"]}
                             for k, v in weights.items()}},
                os.path.join(C.OUT, "problem1_minmax.json"))
    charts(rows, usable)

    # ---- 控制台摘要 ----
    print("===== 问题1 评价结果 =====")
    for r in rows:
        if r.get("image_only"):
            print(f"{r['name']} | 图像型论文（无文字层，已排除）")
        elif r.get("topic") == "人工审核":
            print(f"{r['name']} | 人工审核")
        else:
            print(f"{r['name']} | 赛题{r['topic']} | 得分 {r['score']:.2f} | {r['level']}")
    from collections import Counter
    cnt = Counter(r.get("level", "图像型论文") for r in rows)
    print("\n等级分布:", dict(cnt))
    return rows, usable

# ---------------- 导出 ----------------
def _style_header(ws, headers, widths):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(*(Side(style="thin"),) * 4)
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h); c.font = hf; c.fill = hfill; c.alignment = ca; c.border = thin
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.freeze_panes = "A2"

def _cell(ws, r, c, v, ca=None, fill=None):
    cell = ws.cell(r, c, v)
    if ca: cell.alignment = ca
    cell.border = Border(*(Side(style="thin"),) * 4)
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    return cell

def export_excel(rows, usable, weights, path):
    wb = Workbook()
    ca = Alignment(horizontal="center", vertical="center")
    ws = wb.active; ws.title = "评分结果"
    ind_names = [C.INDICATOR_META[i][0] for i in IND_ORDER]
    headers = ["序号", "论文名称", "赛题类型"] + ind_names + \
              ["U1逻辑严密性", "U2方法合理性", "U3公式推导规范性", "U4结构与文本规范"] + \
              ["不及格隶属度", "及格隶属度", "中等隶属度", "良好隶属度", "优秀隶属度", "综合得分", "综合评价等级"]
    widths = [6, 14, 12] + [14]*16 + [14]*4 + [12]*5 + [10, 14]
    _style_header(ws, headers, widths)
    lv_fill = {"优秀": "C6EFCE", "良好": "FFEB9C", "中等": "BDD7EE", "及格": "F4B084", "不及格": "FFC7CE"}
    r_i = 2
    for r in rows:
        _cell(ws, r_i, 1, r_i-1, ca); _cell(ws, r_i, 2, r["name"], ca)
        if r.get("image_only"):
            _cell(ws, r_i, 3, "图像型论文", ca)
            for c in range(4, len(headers)+1):
                _cell(ws, r_i, c, "—", ca)
            r_i += 1
            continue
        _cell(ws, r_i, 3, r["topic"], ca)
        for k, ind in enumerate(IND_ORDER):
            _cell(ws, r_i, 4+k, round(r["s"][ind], 4), ca)
        for k, dim in enumerate(C.U_ORDER):
            _cell(ws, r_i, 20+k, round(r["dim_scores"][dim], 2), ca)
        for k in range(5):
            _cell(ws, r_i, 24+k, round(float(r["B"][k]), 4), ca)
        _cell(ws, r_i, 29, round(r["score"], 2), ca)
        _cell(ws, r_i, 30, r["level"], ca, lv_fill.get(r["level"]))
        r_i += 1
    ws2 = wb.create_sheet("AHP权重与一致性")
    hdr2 = ["矩阵", "层级", "权重向量", "λmax", "CI", "CR", "是否一致(CR<0.1)"]
    _style_header(ws2, hdr2, [22, 10, 42, 10, 10, 10, 18])
    wrow = 2
    for k, v in weights.items():
        _cell(ws2, wrow, 1, k, ca); _cell(ws2, wrow, 2, "一级" if k == "M(一级)" else "二级", ca)
        _cell(ws2, wrow, 3, "[" + ", ".join(f"{x:.4f}" for x in v["w"]) + "]", ca)
        _cell(ws2, wrow, 4, round(v["lam"], 4), ca); _cell(ws2, wrow, 5, round(v["CI"], 4), ca)
        _cell(ws2, wrow, 6, round(v["CR"], 4), ca); _cell(ws2, wrow, 7, "是" if v["CR"] < 0.1 else "否", ca)
        wrow += 1
    ws3 = wb.create_sheet("等级分布")
    from collections import Counter
    cnt = Counter(r.get("level", "图像型论文") for r in rows)
    _style_header(ws3, ["等级", "篇数", "占比"], [16, 10, 10])
    r3 = 2
    for lv in ["优秀", "良好", "中等", "及格", "不及格", "图像型论文"]:
        _cell(ws3, r3, 1, lv, ca); _cell(ws3, r3, 2, cnt.get(lv, 0), ca)
        _cell(ws3, r3, 3, round(cnt.get(lv, 0)/len(rows), 4), ca); r3 += 1
    wb.save(path)
    print(f"已导出: {path}")

def export_csv(usable, path):
    rows_out = []
    for r in usable:
        rows_out.append({"论文名称": r["name"], "赛题类型": r["topic"],
                         **{C.INDICATOR_META[i][0]: round(r["s"][i], 4) for i in IND_ORDER},
                         **{f"维度_{C.DIM_NAMES[d]}": round(r["dim_scores"][d], 2) for d in C.U_ORDER},
                         **{g: round(float(r["B"][k]), 4) for k, g in enumerate(C.GRADE_NAMES)},
                         "综合得分": round(r["score"], 2), "等级": r["level"]})
    pd.DataFrame(rows_out).to_csv(path, index=False, encoding="utf-8-sig")
    print(f"已导出: {path}")

def charts(rows, usable):
    from collections import Counter
    cnt = Counter(r.get("level", "图像型论文") for r in rows)
    labels = ["优秀", "良好", "中等", "及格", "不及格", "图像型论文"]
    vals = [cnt.get(l, 0) for l in labels]
    C.bar_chart(labels, vals, os.path.join(C.FIGS, "p1_grade_dist.png"),
                title="问题1 三十篇论文质量等级分布", ylabel="篇数", fmt="{:.0f}")
    picks = []
    for key in ["17.pdf", "24.pdf", "30.pdf", "12.pdf"]:
        for r in usable:
            if r["name"] == key:
                picks.append(r)
    if picks:
        C.grouped_bar([C.DIM_NAMES[d] for d in C.U_ORDER],
                      [(r["name"], [round(r["dim_scores"][d], 2) for d in C.U_ORDER]) for r in picks],
                      os.path.join(C.FIGS, "p1_dim_compare.png"),
                      title="代表论文各维度得分对比", ylabel="维度得分", fmt="{:.1f}")
    top = max(usable, key=lambda r: r["score"])
    bot = min(usable, key=lambda r: r["score"])
    cats = [C.INDICATOR_META[i][0] for i in IND_ORDER]
    C.radar_chart(cats, [(top["name"], [top["s"][i] for i in IND_ORDER]),
                         (bot["name"], [bot["s"][i] for i in IND_ORDER])],
                  os.path.join(C.FIGS, "p1_radar_top_bottom.png"),
                  title="最优与最差论文二级指标雷达对比")
    print("图表已生成至", C.FIGS)

if __name__ == "__main__":
    run()