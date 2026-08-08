# -*- coding: utf-8 -*-
"""问题3：针对附件3的3篇“中等”质量论文，设计优化策略。
内容：
1) 基线评估：用问题1模型（参照问题1校准参数）对3篇论文评分，输出二级指标得分；
2) AI生成痕迹检测（离线文本统计特征法）：句长/段落均匀度、套话密度、连接词密度、
   具体性（数字密度）、重复率等合成 AI辅助指数，映射 低/中/高；
3) 逻辑断层识别：规则扫描 模型未推导、缺假设、结论未验证、缺灵敏度/稳健性、
   摘要要素不全、参考文献不足、公式编号/符号说明不足 等；
4) 具体修改方案：按薄弱指标与检测问题给出可操作建议与预计增益；
5) 优化后得分预测：模拟修改后的指标增益，重跑评分模型，输出 优化前→优化后 得分/等级
   及各指标边际贡献。
"""
import os
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import common as C
import problem1
import quantify

TEMPLATE = {
    "u11": "在问题分析、模型推导与结论处补充“因此、综上、进而、由此可得”等逻辑连接词，增强论证衔接",
    "u12": "在摘要与各问题小节中强化对赛题关键词（康养/资源/健康等）的呼应与覆盖",
    "u13": "增加模型检验环节：误差分析、灵敏度分析、收敛性验证与结果吻合性说明",
    "u14": "检查并补全跳过的推导步骤，消除“直接给出结果”的断层",
    "u21": "明确说明所选模型与问题类型的匹配性（如优化/预测/评价问题对应的方法）",
    "u22": "在建模前显式列出基本假设与前提条件，并说明其合理性",
    "u23": "增加与基线模型/对比方案的比较（如线性回归、BP、PSO、贪心等）及优劣分析",
    "u24": "说明方法创新点与相对已有方法的改进之处（如引入XX机制/改进目标函数）",
    "u31": "将核心模型用公式化表达并编号，避免大段文字描述代替公式",
    "u32": "为关键公式添加(1)(2)…编号，并补充符号说明表",
    "u33": "补充公式推导过程与求解步骤（含参数求解、迭代流程）",
    "u34": "统一并集中定义全文符号，增加符号说明/变量定义",
    "u41": "完善标准章节：摘要、引言、问题分析、模型、结果、结论、参考文献齐全",
    "u42": "摘要补全“目的—方法—结果—结论”四要素",
    "u43": "规范引用[1]~[n]，补充并格式统一参考文献",
    "u44": "删减重复赘述，压缩冗余表达，提高信息密度",
}


# ---- 参考论文 AI 检测（5维：3篇内 min-max + 附件1参照偏离）----
AI_WEIGHTS = {"句子重复度": 0.30, "句式变化性": 0.25, "段落均匀度": 0.20,
              "逻辑异常度": 0.15, "术语偏离度": 0.10}
AI_LEVELS = [(0.70, "高"), (0.30, "中")]     # >=0.70 高，>=0.30 中，否则低
EXTRA_AI_BOOST = 0.10                 # 高AI论文：AI相关指标额外提升（降低AI痕迹）
AI_BOOST_INDICATORS = ["u11", "u13", "u44"]   # 逻辑连接、论证闭环、文本冗余

# 参考论文 p1_extract_features.py 的连接词词典（口径一致）
REF_LOGICAL_CONNECTORS = [
    "因此", "所以", "因而", "从而", "于是", "故", "因为", "由于", "基于", "鉴于",
    "如果", "若", "假设", "假定", "设", "则", "那么", "即", "也就是说", "换言之",
    "首先", "其次", "再次", "最后", "然后", "接着", "一方面", "另一方面", "此外", "另外",
    "同时", "然而", "但是", "不过", "尽管", "虽然", "显然", "可见", "由此", "综上所述",
    "总之", "根据", "依照", "按照", "不仅", "而且", "并且", "以及", "换言之",
    "确切地说", "具体而言", "同理", "类似地", "相应地", "反之", "相反", "相对而言",
    "特别地", "尤其", "尤其是",
]
_REF_TERM = re.compile(r"[一-鿿]{2,}")


def ref_ai_raw(body_full):
    """按参考论文公式计算单篇 5 个 AI 原始量（body_full 为剔除代码页后的正文）。"""
    text = body_full
    # 1) 句首重复率：句首 8 字重复占比
    starts = [s.strip()[:8] for s in re.split(r"[。！？\n]+", text) if len(s.strip()) > 10]
    if len(starts) < 2:
        rep_ratio = 0.0
    else:
        sc = {}
        for s in starts:
            sc[s] = sc.get(s, 0) + 1
        repeated = sum(v - 1 for v in sc.values() if v > 1)
        rep_ratio = repeated / len(starts)
    # 2) 句式变化性（burstiness = 句长标准差/均值）
    s_lens = [len(s) for s in re.split(r"[。！？\n]+", text) if len(s.strip()) > 3]
    if len(s_lens) < 2:
        burstiness = 0.0
    else:
        m = np.mean(s_lens)
        burstiness = float(np.std(s_lens) / m) if m > 0 else 0.0
    # 3) 段落均匀度（页块长度变异系数）
    p_lens = [len(p) for p in text.split("\n\n") if len(p.strip()) > 20]
    if len(p_lens) < 2:
        para_uniformity = 0.0
    else:
        m = np.mean(p_lens)
        para_uniformity = float(np.std(p_lens) / m) if m > 0 else 0.0
    # 4) 逻辑连接词密度（每千中文字）
    chinese = len(re.findall(r"[一-鿿]", text))
    connectors = sum(text.count(w) for w in REF_LOGICAL_CONNECTORS)
    logical_density = connectors / (chinese / 1000) if chinese > 0 else 0.0
    # 5) 术语丰富度（中文二字串 unique/total）
    cw = _REF_TERM.findall(text)
    term_rich = len(set(cw)) / len(cw) if cw else 0.0
    return {"rep_ratio": rep_ratio, "burstiness": burstiness, "para_uniformity": para_uniformity,
            "logical_density": logical_density, "term_rich": term_rich}


def ref_ai_reference_stats():
    """附件1参照均值（逻辑连接词密度/术语丰富度，正文口径）；缓存 output/ai_reference_ref.json。"""
    cache = os.path.join(C.OUT, "ai_reference_ref.json")
    if os.path.exists(cache):
        return C.load_json(cache)
    ld, tr = [], []
    for name in sorted(os.listdir(C.ATT1)):
        if not name.lower().endswith(".pdf"):
            continue
        pages, full = C.extract_pdf(os.path.join(C.ATT1, name))
        if C.is_image_only(full):
            continue
        body_pages, body_full = quantify.strip_code_pages(pages)
        raw = ref_ai_raw("\n\n".join(body_pages))
        ld.append(raw["logical_density"])
        tr.append(raw["term_rich"])
    ref = {"logical_density_mean": float(np.mean(ld)), "term_rich_mean": float(np.mean(tr)), "n": len(ld)}
    C.save_json(ref, cache)
    return ref


def ref_ai_index(raw_list, ref):
    """参考论文：5 指标（含相对附件1参照偏离），3 篇内 min-max 归一后加权。"""
    def arr(key):
        return np.array([r[key] for r in raw_list], float)
    ld, tr = arr("logical_density"), arr("term_rich")
    indicators = {
        "句子重复度": arr("rep_ratio"),
        "句式变化性": 1.0 / (np.abs(arr("burstiness") - 0.7) + 0.3),
        "段落均匀度": arr("para_uniformity"),
        "逻辑异常度": np.abs(ld - ref["logical_density_mean"]) / (ref["logical_density_mean"] + 1e-10),
        "术语偏离度": np.abs(tr - ref["term_rich_mean"]) / (ref["term_rich_mean"] + 1e-10),
    }
    norm = {}
    for k, vals in indicators.items():
        vmin, vmax = vals.min(), vals.max()
        norm[k] = (vals - vmin) / (vmax - vmin + 1e-10)
    idx = np.zeros(len(raw_list))
    for k, w in AI_WEIGHTS.items():
        idx += w * norm[k]
    return norm, idx


def ai_level_of(idx):
    for thr, lv in AI_LEVELS:
        if idx >= thr:
            return lv
    return "低"


def detect_issues(t, s_scores, name):
    issues = []
    clean = t["clean"]; full = t["full"]
    if "模型" in clean and t["formula_count"] < 3:
        issues.append(("模型建立", f"模型以文字描述为主，全文公式仅{t['formula_count']}个，公式化不足", "u31/u33"))
    if "模型" in clean and "假设" not in clean:
        issues.append(("模型建立", "未显式给出模型假设与前提条件", "u22"))
    if "结论" in clean or "小结" in clean:
        zone = clean[clean.rfind("结论") if "结论" in clean else clean.rfind("小结"):]
        if not any(w in zone for w in ["验证", "检验", "结果表明", "误差", "灵敏度", "对比", "分析"]):
            issues.append(("结论", "结论部分未回扣结果验证/误差/灵敏度说明", "u13"))
    stab = [w for w in ["灵敏度", "稳健", "稳定性检验", "误差分析", "收敛", "模型检验"] if w in clean]
    if not stab:
        issues.append(("全文", "缺少灵敏度/稳健性/收敛性等模型稳定性分析", "u13"))
    elif not any(w in clean for w in ["灵敏度", "稳健", "误差分析"]):
        issues.append(("全文", "已有收敛性分析，建议补充灵敏度/误差分析以强化稳定性论证", "u13"))
    if "摘要" in clean:
        zone = clean[:clean.find("引言")] if "引言" in clean else clean[:5000]
        has_res = any(w in zone for w in ["结果", "显示", "表明", "得到"])
        has_con = any(w in zone for w in ["结论", "建议", "方案", "提出", "优化"])
        if not (has_res and has_con):
            issues.append(("摘要", "摘要缺少明确的“结果—结论”要素表述", "u42"))
    if t["ref_count"] < 5:
        issues.append(("参考文献", f"参考文献偏少（{t['ref_count']}条）", "u43"))
    if t["formula_count"] >= 3 and t["labeled_formula_count"] < max(2, int(t["formula_count"] * 0.3)):
        issues.append(("公式", "关键公式缺少(1)(2)…编号或符号说明", "u32"))
    if "符号说明" not in clean and "变量定义" not in clean and "符号定义" not in clean:
        issues.append(("全文", "缺少符号说明/变量定义", "u34"))
    return issues

def uplift_for(s):
    if s < 0.30: return 0.30
    if s < 0.50: return 0.22
    if s < 0.70: return 0.12
    return 0.05

def revision_plan(name, s_scores, issues, t):
    """按薄弱指标 + 检测问题生成修改方案。"""
    plan = []
    covered = set()
    weak = [(ind, s) for ind, s in s_scores.items() if s < 0.6]
    for ind, s in sorted(weak, key=lambda x: x[1]):
        covered.add(ind)
        plan.append({"论文": name, "薄弱指标": C.INDICATOR_META[ind][0], "指标编码": ind,
                     "当前得分": round(s, 3), "问题定位": "指标得分低于0.6（" + C.INDICATOR_META[ind][1] + "维度）",
                     "具体修改建议": TEMPLATE[ind], "预计改进后得分": round(min(1.0, s + uplift_for(s)), 3)})
    for loc, desc, inds in issues:
        ind = inds.split("/")[0]
        if ind in covered or inds == "u31/u33":
            pass
        for ii in inds.split("/"):
            if ii not in covered:
                covered.add(ii)
                plan.append({"论文": name, "薄弱指标": C.INDICATOR_META[ii][0], "指标编码": ii,
                             "当前得分": round(s_scores[ii], 3),
                             "问题定位": f"检测发现（{loc}）：{desc}",
                             "具体修改建议": TEMPLATE[ii],
                             "预计改进后得分": round(min(1.0, s_scores[ii] + uplift_for(s_scores[ii])), 3)})
    return plan

def optimize(name, s_scores, plan):
    new_s = dict(s_scores)
    for p in plan:
        new_s[p["指标编码"]] = min(1.0, p["预计改进后得分"])
    return new_s

def run():
    calib = C.load_json(os.path.join(C.OUT, "problem1_minmax.json"))
    refs = {k: v["ref"] for k, v in calib["refs"].items()}
    kinds = problem1.KIND
    centers = {k: np.array(v) for k, v in calib["centers"].items()}
    widths = calib["widths"]; weights = calib["weights"]
    scorer = C.FCEScorer(weights, centers, widths)
    ind_order = {dim: problem1.dim_inds(dim) for dim in C.U_ORDER}

    results = []
    for f in sorted(os.listdir(C.ATT3)):
        if not f.lower().endswith(".pdf"):
            continue
        path = os.path.join(C.ATT3, f)
        pages, full = C.extract_pdf(path)
        t = C.analyze_text(pages, full)
        topic = C.classify_topic(full)
        if topic == "人工审核":
            topic = "A"
        raw = quantify.quantify_all(t, topic)
        s_scores = {ind: float(C.apply_refs(np.array([raw[ind]]), kinds[ind], refs[ind])[0])
                    for ind in problem1.IND_ORDER}
        norm_scores = {dim: [s_scores[ind] for ind in problem1.dim_inds(dim)] for dim in C.U_ORDER}
        B, score = scorer.score(norm_scores, ind_order)
        dim_scores = {}
        for j, dim in enumerate(C.U_ORDER):
            Rj = np.array([scorer.membership(ind, s_scores[ind]) for ind in problem1.dim_inds(dim)])
            dim_scores[dim] = float((scorer.A[f"A{j+1}"] @ Rj) @ C.GRADE_SCORES)
        level = C.level_by_score(score)
        body_pages, _ = quantify.strip_code_pages(pages)
        ai_raw = ref_ai_raw("\n\n".join(body_pages))
        issues = detect_issues(t, s_scores, f)
        plan = revision_plan(f, s_scores, issues, t)
        results.append({"name": f, "topic": topic, "t": t, "s": s_scores, "dim": dim_scores,
                        "score": score, "level": level, "B": B,
                        "ai_raw": ai_raw, "ai": None, "ai_level": None, "ai_scores": None,
                        "issues": issues, "plan": plan,
                        "new_s": None, "score2": None, "level2": None, "margin": None})

    # 参考算法：5 指标（附件1参照偏离）在 3 篇内 min-max 归一后加权
    _ref = ref_ai_reference_stats()
    _norm, _idx = ref_ai_index([r["ai_raw"] for r in results], _ref)
    for j, r in enumerate(results):
        r["ai_scores"] = {k: float(_norm[k][j]) for k in AI_WEIGHTS}
        r["ai"] = float(_idx[j])
        r["ai_level"] = ai_level_of(r["ai"])
    # 优化后得分预测（放到 AI 判定之后，高AI论文额外提升AI相关指标）
    for r in results:
        new_s = optimize(r["name"], r["s"], r["plan"])
        if r["ai_level"] == "高":
            for ind in AI_BOOST_INDICATORS:
                new_s[ind] = min(1.0, new_s[ind] + EXTRA_AI_BOOST)
        norm2 = {dim: [new_s[ind] for ind in problem1.dim_inds(dim)] for dim in C.U_ORDER}
        B2, score2 = scorer.score(norm2, ind_order)
        level2 = C.level_by_score(score2)
        margin = []
        for p in r["plan"]:
            ind = p["指标编码"]
            tmp = dict(r["s"]); tmp[ind] = min(1.0, p["预计改进后得分"])
            nm = {d: [tmp[i] for i in problem1.dim_inds(d)] for d in C.U_ORDER}
            _, sc = scorer.score(nm, ind_order)
            margin.append({"指标": C.INDICATOR_META[ind][0], "改进量": round(new_s[ind] - r["s"][ind], 3),
                           "得分增益": round(sc - r["score"], 2)})
        margin = sorted(margin, key=lambda x: -x["得分增益"])
        r["new_s"] = new_s; r["score2"] = score2; r["level2"] = level2; r["margin"] = margin

    export_excel(results, os.path.join(C.OUT, "problem3_results.xlsx"))
    charts(results)

    print("===== 问题3 结果 =====")
    for r in results:
        print(f"\n{r['name']} | 基线得分 {r['score']:.2f} ({r['level']}) -> 优化后 {r['score2']:.2f} ({r['level2']})")
        print(f"  AI辅助指数 {r['ai']:.3f} -> {r['ai_level']}")
        print(f"  检测问题 {len(r['issues'])} 条, 修改建议 {len(r['plan'])} 条")
        print("  主要增益:", [(m['指标'], m['得分增益']) for m in r['margin'][:4]])
    return results

def export_excel(results, path):
    wb = Workbook()
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(*(Side(style="thin"),) * 4)
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    def sheet(ws, headers, rows, widths=None, wrap_cols=()):
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h); c.font = hf; c.fill = hfill; c.alignment = ca; c.border = thin
            ws.column_dimensions[get_column_letter(i)].width = (widths or [16]*len(headers))[i-1]
        for r_i, row in enumerate(rows, 2):
            for c_i, v in enumerate(row, 1):
                cell = ws.cell(r_i, c_i, v); cell.alignment = ca; cell.border = thin
                if c_i in wrap_cols:
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws.freeze_panes = "A2"
    # sheet1 基线评分
    ws = wb.active; ws.title = "基线评分"
    ind_names = [C.INDICATOR_META[i][0] for i in problem1.IND_ORDER]
    headers = ["论文"] + ind_names + ["U1逻辑严密性", "U2方法合理性", "U3公式推导规范性", "U4结构与文本规范",
               "综合得分", "等级", "AI辅助指数", "AI辅助程度"]
    rows = []
    for r in results:
        rows.append([r["name"]] + [round(r["s"][i], 3) for i in problem1.IND_ORDER] +
                    [round(r["dim"][d], 2) for d in C.U_ORDER] +
                    [round(r["score"], 2), r["level"], round(r["ai"], 3), r["ai_level"]])
    sheet(ws, headers, rows, [10] + [12]*16 + [14]*4 + [10, 10, 12, 12])
    # sheet2 AI检测
    ws2 = wb.create_sheet("AI痕迹检测")
    ai_headers = ["论文"] + list(AI_WEIGHTS.keys()) + ["AI辅助指数", "辅助程度",
                 "句首重复率", "句式变化性(原始)", "段落均匀度(原始)", "逻辑连接词密度", "术语丰富度"]
    ai_rows = []
    for r in results:
        f = r["ai_raw"]; sc = r["ai_scores"]
        ai_rows.append([r["name"]] + [round(sc[k], 3) for k in AI_WEIGHTS] + [round(r["ai"], 3), r["ai_level"],
                       round(f["rep_ratio"], 4), round(f["burstiness"], 3), round(f["para_uniformity"], 3),
                       round(f["logical_density"], 3), round(f["term_rich"], 4)])
    sheet(ws2, ai_headers, ai_rows, [10] + [12]*5 + [12, 12] + [12, 14, 16, 16, 12])
    # sheet3 问题定位
    ws3 = wb.create_sheet("逻辑断层与问题定位")
    q_rows = []
    for r in results:
        for loc, desc, inds in r["issues"]:
            q_rows.append([r["name"], loc, desc, inds])
    sheet(ws3, ["论文", "定位", "检测问题描述", "涉及指标"], q_rows, [10, 12, 60, 14], wrap_cols=(3,))
    # sheet4 修改方案
    ws4 = wb.create_sheet("具体修改方案")
    p_rows = []
    for r in results:
        for p in r["plan"]:
            p_rows.append([p["论文"], p["指标编码"], p["薄弱指标"], p["当前得分"], p["问题定位"],
                           p["具体修改建议"], p["预计改进后得分"]])
    sheet(ws4, ["论文", "指标编码", "薄弱指标", "当前得分", "问题定位", "具体修改建议", "预计改进后得分"],
          p_rows, [10, 10, 16, 10, 40, 55, 14], wrap_cols=(5, 6))
    # sheet5 优化前后
    ws5 = wb.create_sheet("优化前后对比")
    o_rows = []
    for r in results:
        o_rows.append([r["name"], round(r["score"], 2), r["level"], round(r["score2"], 2), r["level2"],
                       round(r["score2"] - r["score"], 2)])
    sheet(ws5, ["论文", "优化前得分", "优化前等级", "优化后得分", "优化后等级", "得分增益"], o_rows, [10, 12, 12, 12, 12, 12])
    # sheet6 边际贡献
    ws6 = wb.create_sheet("边际贡献")
    m_rows = []
    for r in results:
        for m in r["margin"]:
            m_rows.append([r["name"], m["指标"], m["改进量"], m["得分增益"]])
    sheet(ws6, ["论文", "指标", "得分改进量", "边际得分增益"], m_rows, [10, 18, 12, 12])
    wb.save(path)
    print(f"已导出: {path}")

def charts(results):
    names = [r["name"] for r in results]
    C.bar_chart(names, [round(r["ai"], 3) for r in results],
                os.path.join(C.FIGS, "p3_ai_index.png"),
                title="问题3 三篇论文AI辅助指数（离线文本统计特征）", ylabel="AI辅助指数", fmt="{:.2f}")
    C.grouped_bar(names, [("优化前", [round(r["score"], 1) for r in results]),
                          ("优化后", [round(r["score2"], 1) for r in results])],
                  os.path.join(C.FIGS, "p3_before_after.png"),
                  title="优化前后质量得分对比", ylabel="质量得分", fmt="{:.1f}")
    C.radar_chart([C.INDICATOR_META[i][0] for i in problem1.IND_ORDER],
                  [(r["name"], [r["s"][i] for i in problem1.IND_ORDER]) for r in results],
                  os.path.join(C.FIGS, "p3_radar.png"), title="三篇论文二级指标得分雷达图")

if __name__ == "__main__":
    run()