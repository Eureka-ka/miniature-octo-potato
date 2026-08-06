# -*- coding: utf-8 -*-
"""问题2-统计分析：论文质量与可量化文本特征的关联、关键特征识别、
质量调整因子、基于关键特征的质量预测模型，以及小样本稳定性分析（n=9）。
方法：
- 灰色关联度(GRA)：综合维度/原始特征 vs 质量得分；特征剖面 vs 理想剖面。
- Pearson/Spearman 相关 + Bootstrap 95%置信区间 + 置换检验 p 值。
- 关键特征识别：GRA + 相关显著性 + 语义（逻辑连接、公式规范、参考文献）三重证据。
- 质量预测模型：标准化多元线性回归与岭回归（λ=1.0），LOO-CV 外推评估。
- 小样本稳定性：LOO-CV(R²/RMSE/MAE)、Bootstrap(1000次)系数分布、单样本剔除敏感性。
- 质量调整因子：k_i = 1 + α·w·(ŷ_i-Q_i)/Q_i + β·(F_i-F̄)/F̄，
  其中 w=clip(R²_LOO/0.5,0,1) 为模型外推可靠性权重，F_i 为特征剖面与理想剖面的GRA。
"""
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import common as C
import problem2_features

RNG = np.random.default_rng(42)
RHO = 0.5
ALPHA = 0.3        # 回归调整分量的收缩系数
BETA = 0.1         # 特征剖面分量的收缩系数
LAMBDA = 1.0       # 岭回归正则参数
R2_REF = 0.5       # “可接受外推能力”基准
KEY_FEATURES = ["逻辑连词总频次", "规范编号公式占比", "参考文献数量"]

def pearson(x, y):
    x = np.asarray(x, float); y = np.asarray(y, float)
    if np.std(x) < 1e-12 or np.std(y) < 1e-12:
        return 0.0
    return float(np.corrcoef(x, y)[0, 1])

def spearman(x, y):
    def rank(a):
        order = np.argsort(a); ranks = np.empty_like(a, float)
        ranks[order] = np.arange(1, len(a) + 1)
        return ranks
    return pearson(rank(np.asarray(x)), rank(np.asarray(y)))

def bootstrap_ci(x, y, n=1000, seed=42):
    rng = np.random.default_rng(seed); idx = np.arange(len(x)); rs = []
    for _ in range(n):
        s = rng.choice(idx, size=len(idx), replace=True)
        rs.append(pearson(x[s], y[s]))
    return np.percentile(rs, [2.5, 97.5])

def perm_p(x, y, n=2000, seed=7):
    rng = np.random.default_rng(seed); obs = abs(pearson(x, y)); cnt = 0
    for _ in range(n):
        if abs(pearson(x, rng.permutation(y))) >= obs:
            cnt += 1
    return (cnt + 1) / (n + 1)

def gra_seq(reference, matrix):
    """灰色关联度：reference 序列 vs 每个比较序列（已方向一致归一化）。"""
    ref = np.asarray(reference, float); X = np.asarray(matrix, float)
    delta = np.abs(X - ref[None, :]); dmin = delta.min(); dmax = delta.max()
    if abs(dmax - dmin) < 1e-12:
        return np.ones(X.shape[0])
    xi = (dmin + RHO * dmax) / (delta + RHO * dmax)
    return xi.mean(axis=1)

def gra_to_ideal(profile_matrix):
    """每篇论文特征剖面与理想剖面（各维度最大值）的灰色关联度。"""
    X = np.asarray(profile_matrix, float)
    ref = X.max(axis=0)
    return gra_seq(ref, X)

def standardize(x):
    x = np.asarray(x, float)
    return (x - x.mean(axis=0)) / (x.std(axis=0) + 1e-12)

def ols(X, y):
    X1 = np.column_stack([np.ones(len(y)), X])
    b, *_ = np.linalg.lstsq(X1, y, rcond=None)
    return b

def ridge(X, y, lam=LAMBDA):
    X1 = np.column_stack([np.ones(len(y)), X])
    eye = np.eye(X1.shape[1]); eye[0, 0] = 0.0
    return np.linalg.solve(X1.T @ X1 + lam * eye, X1.T @ y)

def loo_cv(X, y, kind="ridge", lam=LAMBDA):
    n = len(y); preds = []
    for i in range(n):
        tr = [k for k in range(n) if k != i]
        Xtr = standardize(X[tr]); ytr = standardize(y[tr])
        b = (ridge(Xtr, ytr, lam) if kind == "ridge" else ols(Xtr, ytr))
        Xte = (X[i] - X[tr].mean(axis=0)) / (X[tr].std(axis=0) + 1e-12)
        p = float(b[1:] @ Xte + b[0])
        preds.append(p * y[tr].std() + y[tr].mean())
    preds = np.asarray(preds)
    ss_res = float(np.sum((y - preds) ** 2)); ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    rmse = float(np.sqrt(np.mean((y - preds) ** 2)))
    mae = float(np.mean(np.abs(y - preds)))
    return r2, rmse, mae, preds

def metrics(y, yhat):
    y = np.asarray(y); yhat = np.asarray(yhat)
    ss_res = float(np.sum((y - yhat) ** 2)); ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    rmse = float(np.sqrt(np.mean((y - yhat) ** 2)))
    mae = float(np.mean(np.abs(y - yhat)))
    return r2, rmse, mae

def run():
    papers = C.load_json(os.path.join(C.OUT, "problem2_papers.json"))
    n = len(papers)
    names = [p["name"] for p in papers]
    y = np.array([p["score"] for p in papers], float)

    feats_raw = np.array([[p["feats"][f] for f in problem2_features.FEATURES] for p in papers], float)
    feats_norm = feats_raw.copy()
    for j, f in enumerate(problem2_features.FEATURES):
        feats_norm[:, j] = C.zscore_norm(feats_raw[:, j], -1 if f in problem2_features.NEGATIVE else 1)  # 反向指标取负
    dim_names = ["篇幅结构", "公式特征", "逻辑连接", "参考文献"]
    dims = np.column_stack([
        feats_norm[:, [problem2_features.FEATURES.index(c) for c in cols]].mean(axis=1)
        for cols in problem2_features.GROUP_DEF.values()])

    # ---- 灰色关联度（参考序列与特征同尺度：均用 Z-score）----
    y_norm = C.zscore_norm(y)
    gra_dim = gra_seq(y_norm, dims.T)
    gra_raw = gra_seq(y_norm, feats_norm.T)
    F = gra_to_ideal(dims)                      # 特征剖面 vs 理想剖面

    # ---- 相关性 ----
    corr_rows = []
    for j, f in enumerate(problem2_features.FEATURES):
        x = feats_norm[:, j]
        r = pearson(x, y); rho = spearman(x, y)
        ci = bootstrap_ci(x, y); p = perm_p(x, y)
        corr_rows.append({"特征": f, "Pearson_r": round(r, 4), "Spearman_rho": round(rho, 4),
                          "Bootstrap_CI下限": round(ci[0], 4), "Bootstrap_CI上限": round(ci[1], 4),
                          "置换检验p": round(p, 4), "GRA": round(float(gra_raw[j]), 4)})
    corr_df = pd.DataFrame(corr_rows)
    corr_df["abs_r"] = corr_df["Pearson_r"].abs()
    key_df = corr_df.sort_values("abs_r", ascending=False).reset_index(drop=True)

    # ---- 预测模型（关键特征 = 逻辑连接/公式规范/参考文献）----
    X_raw = feats_norm[:, [problem2_features.FEATURES.index(f) for f in KEY_FEATURES]]
    X_std = standardize(X_raw); y_std = standardize(y)
    b_ols = ols(X_std, y_std); b_ridge = ridge(X_std, y_std)
    beta_std = b_ridge[1:]
    beta_orig = beta_std * (y.std() / (X_raw.std(axis=0) + 1e-12))
    intercept_orig = y.mean() - float(beta_orig @ X_raw.mean(axis=0))
    yhat_in = X_std @ beta_std + b_ridge[0]
    r2_in, rmse_in, mae_in = metrics(y_std, yhat_in)

    # ---- LOO-CV（OLS 与 岭回归）----
    r2_loo_o, rmse_loo_o, mae_loo_o, pred_o = loo_cv(X_raw, y, "ols")
    r2_loo, rmse_loo, mae_loo, pred_r = loo_cv(X_raw, y, "ridge")
    loo_df = pd.DataFrame({"论文": names, "实际得分": np.round(y, 2),
                           "OLS预测": np.round(pred_o, 2), "岭回归预测": np.round(pred_r, 2),
                           "残差(岭)": np.round(y - pred_r, 2)})

    # ---- Bootstrap 系数分布（岭回归）----
    bs = []
    for _ in range(1000):
        s = RNG.choice(n, size=n, replace=True)
        Xb = standardize(X_raw[s]); yb = standardize(y[s])
        bs.append(ridge(Xb, yb)[1:])
    bs = np.array(bs)
    bs_df = pd.DataFrame({"特征": KEY_FEATURES, "系数均值": np.round(bs.mean(axis=0), 4),
                          "95%CI下限": np.round(np.percentile(bs, 2.5, axis=0), 4),
                          "95%CI上限": np.round(np.percentile(bs, 97.5, axis=0), 4),
                          "系数标准差": np.round(bs.std(axis=0), 4)})

    # ---- 单样本剔除敏感性（岭回归）----
    sens = []
    for i in range(n):
        tr = [k for k in range(n) if k != i]
        b = ridge(standardize(X_raw[tr]), standardize(y[tr]))[1:]
        rel = np.abs((b - beta_std) / (np.abs(beta_std) + 1e-9)) * 100
        sens.append({"剔除论文": names[i],
                     **{f"β({KEY_FEATURES[j]})变化%": round(float(rel[j]), 1) for j in range(3)},
                     "最大变化%": round(float(rel.max()), 1)})
    sens_df = pd.DataFrame(sens)

    # ---- 质量调整因子 ----
    Q = y; Qpred = pred_r
    w = float(np.clip(r2_loo / R2_REF, 0.0, 1.0))     # 外推可靠性权重
    Fmean = F.mean()
    k = 1 + ALPHA * w * (Qpred - Q) / np.maximum(Q, 1e-9) + BETA * (F - Fmean) / max(Fmean, 1e-9)
    Qadj = Q * k
    adj_df = pd.DataFrame({"论文": names, "基础得分Q": np.round(Q, 2),
                           "特征预测得分(岭)": np.round(Qpred, 2),
                           "特征剖面GRA(F)": np.round(F, 4),
                           "调整因子k": np.round(k, 4),
                           "调整后得分Q_adj": np.round(Qadj, 2),
                           "原等级": [C.level_by_score(q) for q in Q],
                           "调整后等级": [C.level_by_score(q) for q in Qadj]})

    # ---- 导出 ----
    export_excel(corr_df, key_df, gra_dim, dim_names, gra_raw, bs_df, loo_df, sens_df,
                 adj_df, KEY_FEATURES, beta_std, beta_orig, intercept_orig,
                 r2_in, rmse_in, mae_in, r2_loo, rmse_loo, mae_loo,
                 r2_loo_o, rmse_loo_o, mae_loo_o, w, os.path.join(C.OUT, "problem2_stats.xlsx"))
    charts(y, names, corr_df, gra_dim, dim_names, bs_df, KEY_FEATURES, loo_df, Q, Qadj)

    print("===== 问题2 统计结果 =====")
    print("灰色关联度(综合维度):", dict(zip(dim_names, np.round(gra_dim, 4))))
    print("关键特征:", KEY_FEATURES)
    print(key_df[["特征", "Pearson_r", "置换检验p", "GRA"]].to_string(index=False))
    print(f"岭回归 样本内 R2={r2_in:.3f} RMSE={rmse_in:.3f} | LOO-CV R2={r2_loo:.3f} RMSE={rmse_loo:.3f} MAE={mae_loo:.3f}")
    print(f"OLS LOO-CV R2={r2_loo_o:.3f}（小样本不稳定） | 可靠性权重 w={w:.3f}")
    print(adj_df.to_string(index=False))
    return locals()

def export_excel(corr_df, key_df, gra_dim, dim_names, gra_raw, bs_df, loo_df, sens_df,
                 adj_df, keyfeats, beta_std, beta_orig, intercept_orig,
                 r2_in, rmse_in, mae_in, r2_loo, rmse_loo, mae_loo,
                 r2_loo_o, rmse_loo_o, mae_loo_o, w, path):
    wb = Workbook()
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(*(Side(style="thin"),) * 4)
    def sheet(ws, headers, rows, widths=None):
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h); c.font = hf; c.fill = hfill; c.alignment = ca; c.border = thin
            ws.column_dimensions[get_column_letter(i)].width = (widths or [15]*len(headers))[i-1]
        for r_i, row in enumerate(rows, 2):
            for c_i, v in enumerate(row, 1):
                cell = ws.cell(r_i, c_i, v); cell.alignment = ca; cell.border = thin
        ws.freeze_panes = "A2"
    ws1 = wb.active; ws1.title = "灰色关联度-维度"
    sheet(ws1, ["综合维度", "灰色关联度"], [[dn, round(float(g), 4)] for dn, g in zip(dim_names, gra_dim)])
    ws2 = wb.create_sheet("灰色关联度-原始特征")
    feats = problem2_features.FEATURES
    sheet(ws2, ["特征", "灰色关联度"], [[f, round(float(g), 4)] for f, g in zip(feats, gra_raw)])
    ws3 = wb.create_sheet("相关性分析")
    sheet(ws3, list(corr_df.columns), corr_df.values.tolist())
    ws4 = wb.create_sheet("关键特征排序")
    sheet(ws4, list(key_df.columns), key_df.values.tolist())
    ws5 = wb.create_sheet("回归模型")
    rows5 = [["预测特征", "标准化系数(岭)", "原始尺度系数", "截距(原始)"],
             [keyfeats[0], round(float(beta_std[0]), 4), round(float(beta_orig[0]), 4), ""],
             [keyfeats[1], round(float(beta_std[1]), 4), round(float(beta_orig[1]), 4), ""],
             [keyfeats[2], round(float(beta_std[2]), 4), round(float(beta_orig[2]), 4), ""],
             ["截距", "", "", round(float(intercept_orig), 4)],
             ["样本内R²/RMSE/MAE", round(r2_in, 4), round(rmse_in, 4), round(mae_in, 4)],
             ["LOO-CV R²/RMSE/MAE", round(r2_loo, 4), round(rmse_loo, 4), round(mae_loo, 4)],
             ["OLS LOO-CV R²", round(r2_loo_o, 4), "（对比）", "", ""]]
    sheet(ws5, rows5[0], rows5[1:], [24, 18, 18, 14])
    ws6 = wb.create_sheet("LOO交叉验证")
    sheet(ws6, list(loo_df.columns), loo_df.values.tolist())
    ws7 = wb.create_sheet("Bootstrap系数")
    sheet(ws7, list(bs_df.columns), bs_df.values.tolist())
    ws8 = wb.create_sheet("敏感性-单样本剔除")
    sheet(ws8, list(sens_df.columns), sens_df.values.tolist())
    ws9 = wb.create_sheet("质量调整因子")
    sheet(ws9, list(adj_df.columns), adj_df.values.tolist())
    wb.save(path)
    print(f"已导出: {path}")

def charts(y, names, corr_df, gra_dim, dim_names, bs_df, keyfeats, loo_df, Q, Qadj):
    C.bar_chart(dim_names, [round(float(g), 3) for g in gra_dim],
                os.path.join(C.FIGS, "p2_gra_dims.png"),
                title="问题2 各综合维度与质量得分的灰色关联度", ylabel="灰色关联度", fmt="{:.2f}")
    k6 = corr_df.sort_values("abs_r", ascending=False).head(6)
    C.hbar_chart(k6["特征"].tolist(), k6["Pearson_r"].tolist(),
                 os.path.join(C.FIGS, "p2_corr.png"), title="特征与质量得分的Pearson相关(±95%CI)",
                 errors=(k6["Pearson_r"] - k6["Bootstrap_CI下限"]).abs().tolist(),
                 fmt="{:.2f}", xmin=-1, xmax=1)
    bc = bs_df["系数均值"].tolist(); be = (bs_df["系数均值"] - bs_df["95%CI下限"]).abs().tolist()
    m = max([abs(v) + (e or 0) for v, e in zip(bc, be)] + [0.5]) * 1.15
    C.hbar_chart(bs_df["特征"].tolist(), bc, os.path.join(C.FIGS, "p2_coef_bootstrap.png"),
                 title="关键特征标准化系数(岭回归, Bootstrap均值±95%CI)",
                 errors=be, fmt="{:.2f}", xmin=-m, xmax=m)
    C.line_chart(loo_df["论文"].tolist(),
                 [("实际得分", loo_df["实际得分"].values.tolist()),
                  ("岭回归预测", loo_df["岭回归预测"].values.tolist())],
                 os.path.join(C.FIGS, "p2_loo.png"), title="留一交叉验证：实际 vs 预测质量得分",
                 ylabel="得分", fmt="{:.1f}")
    C.grouped_bar(names, [("基础得分Q", Q.tolist()), ("调整后得分Q_adj", Qadj.tolist())],
                  os.path.join(C.FIGS, "p2_adjust.png"), title="质量调整因子作用前后对比",
                  ylabel="得分", fmt="{:.1f}", rotate=45)

if __name__ == "__main__":
    run()