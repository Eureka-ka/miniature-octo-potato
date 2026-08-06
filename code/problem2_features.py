# -*- coding: utf-8 -*-
"""问题2-特征工程：提取10篇同赛题论文的可量化文本特征（12项）。

升级点（与问题1专属量化对齐）：
1) 提取前剔除代码附录页（strip_code_pages），总页数→内容页数、总字符数→内容字符数，
   公式/连词/参考文献/章节等均在正文上计算，避免代码页污染。
2) 章节完整度用行首标题级检测(line_start_keyword)；公式总数用行级公式检测(detect_equation_lines)；
   规范编号公式占比用逐式编号(count_labeled_equations)；参考文献数量用引用-条目(parse_references)。

质量得分仍沿用问题1的AHP-模糊综合评价（参照问题1校准参数，使用全文文本以保持口径不变）。
2-8.pdf为图像型论文，予以排除（n=9）。
"""
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import common as C
import problem1
import quantify

# 行首标题级章节词表（收紧，避免正文“模型：…”误判；供 line_start_keyword 使用）
CHAPTER_GROUPS = [
    ["摘要"],
    ["引言", "绪论", "introduction"],
    ["模型建立", "模型构建", "模型的建立", "建模", "模型求解", "求解方法"],
    ["结果分析", "仿真结果", "实验结果", "结果与分析"],
    ["结论", "小结", "总结", "conclusion"],
    ["参考文献", "references"],
]
FEATURES = ["内容页数", "内容字符数", "章节完整度", "段落平均长度",
            "公式总数", "公式密度(公式数/内容字符)", "规范编号公式占比",
            "逻辑连词总频次", "连词密度(连词数/内容字符)", "参考文献数量"]
GROUP_DEF = {
    "篇幅结构特征综合得分": ["内容页数", "内容字符数", "章节完整度", "段落平均长度"],
    "公式特征综合得分": ["公式总数", "公式密度(公式数/内容字符)", "规范编号公式占比"],
    "逻辑连接特征综合得分": ["逻辑连词总频次", "连词密度(连词数/内容字符)"],
    "参考文献特征综合得分": ["参考文献数量"],
}
NEGATIVE = {"段落平均长度"}

def extract_features(pages, full, t):
    """在“剔除代码页后的正文”上计算12项特征。pages/full 为正文页列表与正文全文。"""
    chapter_hit = sum(1 for grp in CHAPTER_GROUPS if quantify.line_start_keyword(full, grp))
    eq_idx, _ = quantify.detect_equation_lines(full)
    n_eq = len(eq_idx)
    feats = {
        "内容页数": len(pages),
        "内容字符数": t["total_chars"],
        "章节完整度": round(chapter_hit / len(CHAPTER_GROUPS), 3),
        "段落平均长度": round(t["total_chars"] / max(1, len(t["paragraphs"])), 2),
        "公式总数": n_eq,
        "公式密度(公式数/内容字符)": round(n_eq / max(1, t["total_chars"]), 6),
        "规范编号公式占比": round(quantify.count_labeled_equations(full, eq_idx) / max(1, n_eq), 3),
        "逻辑连词总频次": t["connector_count"],
        "连词密度(连词数/内容字符)": round(t["connector_count"] / max(1, t["total_chars"]), 6),
        "参考文献数量": quantify.parse_references(full)[1],
    }
    return feats

def norm_features(df):
    df = df.copy()
    for col in FEATURES:
        arr = df[col].astype(float).values
        if col == "章节完整度":
            continue  # 本身0~1
        s = C.minmax_norm(arr, -1 if col in NEGATIVE else 1)
        df[col + "_归一"] = np.round(s, 6)
    for gname, cols in GROUP_DEF.items():
        df[gname] = df[[c + "_归一" if c != "章节完整度" else c for c in cols]].mean(axis=1)
    return df

def run():
    files = sorted(f for f in os.listdir(C.ATT2) if f.lower().endswith(".pdf"))
    rows = []
    for f in files:
        path = os.path.join(C.ATT2, f)
        pages, full = C.extract_pdf(path)
        img = C.is_image_only(full)
        base = {"论文名称": f, "数据来源": "图像型论文(排除)" if img else "自动识别"}
        if img:
            rows.append(base)
            continue
        # 质量得分用全文（保持口径不变）；特征用剔除代码页后的正文
        t_full = C.analyze_text(pages, full)
        body_pages, body_full = quantify.strip_code_pages(pages)
        t_body = C.analyze_text(body_pages, body_full)
        feats = extract_features(body_pages, body_full, t_body)
        topic = C.classify_topic(full)
        raw = quantify.quantify_all(t_full, topic if topic != "人工审核" else "B")
        rows.append(dict(base, feats=feats, raw=raw, topic=topic,
                         code_pages=len(pages) - len(body_pages), total_pages=len(pages)))

    usable = [r for r in rows if r.get("feats") is not None]
    df = pd.DataFrame([{"论文名称": r["论文名称"]} | r["feats"] for r in usable])

    # ---- 综合质量得分（沿用问题1校准参数）----
    calib = C.load_json(os.path.join(C.OUT, "problem1_minmax.json"))
    refs = {k: v["ref"] for k, v in calib["refs"].items()}
    kinds = problem1.KIND
    centers = {k: np.array(v) for k, v in calib["centers"].items()}
    widths = calib["widths"]
    weights = calib["weights"]
    scorer = C.FCEScorer(weights, centers, widths)
    ind_order = {dim: problem1.dim_inds(dim) for dim in C.U_ORDER}
    scores = []
    for r in usable:
        s = {ind: float(C.apply_refs(np.array([r["raw"][ind]]), kinds[ind], refs[ind])[0]) for ind in problem1.IND_ORDER}
        norm_scores = {dim: [s[ind] for ind in problem1.dim_inds(dim)] for dim in C.U_ORDER}
        B, score = scorer.score(norm_scores, ind_order)
        scores.append({"B": B, "score": score, "level": C.level_by_score(score)})
    for r, sc in zip(usable, scores):
        r["score"] = sc["score"]; r["level"] = sc["level"]; r["B"] = sc["B"]
    df["综合质量得分"] = [round(r["score"], 2) for r in usable]
    df["等级"] = [r["level"] for r in usable]
    df["赛题类型"] = [r["topic"] for r in usable]

    df_norm = norm_features(df)
    out_cols = ["论文名称", "数据来源", "赛题类型"] + FEATURES + \
               [c + "_归一" for c in FEATURES if c != "章节完整度"] + \
               list(GROUP_DEF.keys()) + ["综合质量得分", "等级"]
    df_norm["数据来源"] = "自动识别"
    df_norm["赛题类型"] = df["赛题类型"]
    df_out = df_norm[out_cols]
    df_out.to_csv(os.path.join(C.OUT, "problem2_features.csv"), index=False, encoding="utf-8-sig")
    export_excel(df_out, os.path.join(C.OUT, "problem2_features.xlsx"))
    # 保存供统计脚本使用（含代码页数供报告）
    C.save_json([{"name": r["论文名称"], "feats": r["feats"], "score": r["score"],
                  "level": r["level"], "topic": r["topic"],
                  "code_pages": r["code_pages"], "total_pages": r["total_pages"]} for r in usable],
                os.path.join(C.OUT, "problem2_papers.json"))
    print("===== 问题2 特征与质量得分（已剔除代码页）=====")
    for r in usable:
        print(f"{r['论文名称']} | 总页数{r['total_pages']} 代码页{r['code_pages']} 内容页数{r['feats']['内容页数']} | 得分 {r['score']:.2f} | {r['level']}")
    return usable, df_out

def export_excel(df, path):
    wb = Workbook()
    ws = wb.active; ws.title = "特征与质量得分"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(*(Side(style="thin"),) * 4)
    headers = list(df.columns)
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h); c.font = hf; c.fill = hfill; c.alignment = ca; c.border = thin
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(h)) * 2)
    for r_i, row in enumerate(df.itertuples(index=False), 2):
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(r_i, c_i, v)
            cell.alignment = ca; cell.border = thin
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"已导出: {path}")

if __name__ == "__main__":
    run()